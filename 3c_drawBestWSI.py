import numpy as np
import configparser
import argparse
import os
import time
import openslide
import cv2
from utils import print_config
import sys
import openpyxl





# =============================================================================
# FUNCTIONS
# =============================================================================
def findGTLineFromDetectionLine(line_detections, lines_gts):
    for line_gts in lines_gts :
        if line_detections.split(' ')[0].split('/')[-1] == line_gts.split(' ')[0].split('/')[-1] :
                return line_gts





def drawRectangleInImage(x1, y1, x2, y2, img, color, th):
    cv2.rectangle(img, (x1, y1), (x2, y2), color, th)
    




def convertDetections(detections):
    return detections.replace('\'','').replace('"','').replace('[','').replace(']','').replace(' ','').replace('\n','').split('glomeruli')[1:]

    

    

def result_is_a_correct_detection(gt, detection, threshold_iou):
    if len(gt) == 0: 
         return False    
     
    # create table filled with 0 (no dectection), 1 (predicted detection), 2 (groundtruth detection) and 3 (1 & 2)
    result_pixels = np.array([[0 for col in range(1500)] for row in range(1500)])
    
    for x in range(int(detection[0]), int(detection[2])):
        for y in range(int(detection[1]), int(detection[3])):
            result_pixels[x, y] = 1
    
    gt_pixels = np.array([[0 for col in range(1500)] for row in range(1500)])
    
    for x in range(int(gt[0]), int(gt[2])):
        for y in range(int(gt[1]), int(gt[3])):
            gt_pixels[x, y] = 2
                
    merge_pixels = result_pixels + gt_pixels
    
    # Compute the shared area between the groundtruth and predicted detections
    number_of_pixels_detected_by_prediction = np.count_nonzero(merge_pixels == 1)
    number_of_pixels_detected_by_gt = np.count_nonzero(merge_pixels == 2)
    number_of_pixels_detected_by_gt_and_prediction = np.count_nonzero(merge_pixels == 3)
    
    shared_area = number_of_pixels_detected_by_gt_and_prediction / (number_of_pixels_detected_by_prediction + number_of_pixels_detected_by_gt + number_of_pixels_detected_by_gt_and_prediction)

    if shared_area > threshold_iou:
        return True
        
    return False





def build_confusion_matrice_for_one_image(gt, result, threshold_iou):
    TP = []
    FP = []
    FN = []

    # formating data
    formating_result = []
    result = result.split('.png ')[1:]
    for detection in result[0].replace('[[', '').replace(']]', '').split('], ['):
        positions = detection.split(', ')
        positions.append(positions.pop(0))
        
        formating_result.append(positions)
    
    formating_gt = []

    gt = gt.split('.png ')[1:]

    for detection in gt[0].split(' '):
        positions = detection.split(',')
        positions.pop()
        
        formating_gt.append(positions)
    
    # classify detections
    # filter with confidence score
    new_formating_result = []
    for result in formating_result:
        if float(result[-1].split(' ')[-1][:-1]) >= confidence:
            new_formating_result.append(result)
            
    formating_result = new_formating_result
    # filter with IOU
    for gt in formating_gt:
        detection_is_correct = False

        for result in formating_result:
            if result_is_a_correct_detection(gt, result, threshold_iou):
                detection_is_correct = True
                TP.append(result)

        if not detection_is_correct:
            FN.append(gt)
    # If a detection is not correct (TP), the detection is FP.
    for r in formating_result:
        if r not in TP:
            FP.append(r)
    
    return TP, FP, FN





# =============================================================================
# MAIN
# =============================================================================
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Draw WSI with detections for IOU and Confidence score with higher F1Score.')
    parser.add_argument('-c', '--config', type=str, help='The configuration file to use')
    args = parser.parse_args()
    
    # Read the config file given in parameter, if not read the default config file
    if args.config:
        config_path = args.config
    else:
        config_path = 'config.cfg'
        
    if not os.path.isfile(config_path):
        raise ValueError('The configuration file path given does not exist : ', config_path)
        
    config = configparser.RawConfigParser(interpolation = configparser.ExtendedInterpolation())
    config.read(config_path)
    
    
    display_logs = config.getboolean('General', 'display_logs')
    path_logs = config.get('Paths', 'path_logs')

    patients_test = config.get('Train', 'patients_test')
    path_images = config.get('Paths', 'path_images')
    lod = config.getint('MultiScale', 'lod')
    scale_list = config.get('MultiScale', 'scale_list')
    tmp = []
    for e in scale_list.split(','):
        tmp.append(int(e))
    scale_list = tmp

    path_dataset_created = config.get('Paths', 'path_dataset_created')
    dataset_name = config.get('Train', 'dataset_name')
    annotations_file_name = config.get('Paths', 'final_annotations_file_name')
    annotation_path = path_dataset_created + dataset_name + '/' + annotations_file_name
    
    
    
    
    sys.stdout = open(path_logs + '/log_DrawBestWSI.txt', 'w')
    
    if display_logs :
        print_config(config_path)
    
    # Open annotations file
    with open(annotation_path) as f_gts:
        lines_gts = f_gts.readlines()
        
        
        
    # Open results file
    annotations_results_file_name = path_logs + '/annotations_results.txt'
    
    with open(annotations_results_file_name) as f_detections:
        lines_detections = f_detections.readlines()
        
    

    wrapper_lines_detections = []
    for s in scale_list:
        tmp = []
        for line_detections in lines_detections :
            if ('scale_' + str(s)) in line_detections.split(' ')[0].split('/')[-1] :
                tmp.append(line_detections)
        wrapper_lines_detections.append(tmp)
    
    start = time.time()    
    
    for lines_detections in wrapper_lines_detections:
        scale = int(lines_detections[0].split(' ')[0].split('/')[-1].split('_')[-1][:-4])
        
        results = []

        # find best IOU & Confidence 
        workbook = openpyxl.load_workbook(path_logs + '/datas_' + str(scale) + '.xlsx')
        worksheet = workbook.active 

        max_column = worksheet.max_column 
        max_row = worksheet.max_row 

        results = []
        result = []

        for row in range(1, max_row): 
            result = []
            for column in range(0, max_column):
                result.append(worksheet.cell(row=row+1, column=column+1).value)
            results.append(result)

        workbook.close()

        best_F1Score = 0
        iou = 0 
        confidence = 0
        for r in results:
            if r[7] > best_F1Score:
                best_F1Score = r[7]
                iou = r[0]
                confidence = r[1]




        print('\nEvaluation in progress :\tIOU', iou,'\tConfidence', confidence)

        for patient in patients_test.split(',') :

            img = openslide.OpenSlide(path_images + '/IFTA_00' + str(patient) + '_02.svs')

            img_patch = img.read_region((100, 100), 1,(img.dimensions[0]//(lod*lod) ,img.dimensions[1]//(lod*lod)))

            img_patch_resized = cv2.resize(np.array(img_patch), (img_patch.size[0], img_patch.size[1]))

            tp_count = 0
            fp_count = 0
            fn_count = 0

            for line_detections in lines_detections :
                TP = []
                FP = []
                FN = []

                if ('IFTA_00' + patient) in line_detections.split(' ')[0].split('/')[-1] :
                    line_gts = findGTLineFromDetectionLine(line_detections, lines_gts)

                    detections = convertDetections(line_detections)
                    gts = line_gts.split(' ')[1:]

                    patch_x = int(line_detections.split(' ')[0].split('/')[-1].split('_')[-5]) - int(100/(lod*lod))
                    patch_y = int(line_detections.split(' ')[0].split('/')[-1].split('_')[-3]) - int(100/(lod*lod))

                    TP, FP, FN = build_confusion_matrice_for_one_image(line_gts, line_detections, iou)

                    for tp in TP:
                        annotation_x1 = int((int(tp[0]) * scale)/(lod*lod))
                        annotation_y1 = int((int(tp[1]) * scale)/(lod*lod))
                        annotation_x2 = int((int(tp[2]) * scale)/(lod*lod))
                        annotation_y2 = int((int(tp[3]) * scale)/(lod*lod))
                        drawRectangleInImage(patch_x + annotation_x1, patch_y + annotation_y1, patch_x + annotation_x2, patch_y + annotation_y2,  img_patch_resized, (0, 255, 0),20)

                    for fp in FP:
                        annotation_x1 = int((int(fp[0]) * scale)/(lod*lod))
                        annotation_y1 = int((int(fp[1]) * scale)/(lod*lod))
                        annotation_x2 = int((int(fp[2]) * scale)/(lod*lod))
                        annotation_y2 = int((int(fp[3]) * scale)/(lod*lod))
                        drawRectangleInImage(patch_x + annotation_x1, patch_y + annotation_y1, patch_x + annotation_x2, patch_y + annotation_y2,  img_patch_resized, (255, 0, 0),20)

                    for fn in FN:
                        annotation_x1 = int((int(fn[0]) * scale)/(lod*lod))
                        annotation_y1 = int((int(fn[1]) * scale)/(lod*lod))
                        annotation_x2 = int((int(fn[2]) * scale)/(lod*lod))
                        annotation_y2 = int((int(fn[3]) * scale)/(lod*lod))
                        drawRectangleInImage(patch_x + annotation_x1, patch_y + annotation_y1, patch_x + annotation_x2, patch_y + annotation_y2,  img_patch_resized, (0, 0, 255),20)

                tp_count += len(TP)
                fp_count += len(FP)
                fn_count += len(FN)

            img_patch_resized = cv2.resize(img_patch_resized, (len(img_patch_resized[0])//2, len(img_patch_resized)//2))            
            cv2.imwrite((path_logs + '/ResultsOnWSI_patient_' + str(patient) + '_scale_' + str(scale) + '_IOU_' + str(iou) + '_confidenceScore_' + str(confidence) + '.png'), cv2.cvtColor(img_patch_resized, cv2.COLOR_RGB2BGR))
            print('Image created for patient', patient, '- scale', str(scale))

            precision = tp_count / (tp_count+fp_count)
            rappel = tp_count / (tp_count+fn_count)
            F1score = (2*precision*rappel) / (precision+rappel)

            print('TP\t\t', tp_count, '\nFP\t\t', fp_count, '\nFN\t\t', fn_count)
            print('precision\t', round(precision,3), '\nrappel\t\t', round(rappel,3))
            print('F1 Score\t', round(F1score,3))

            results.append([iou, confidence, tp_count, fp_count, fn_count, precision, rappel, F1score, patient])

    end = time.time()
    print('Execution time :', round(((end-start)/60), 2), 'min')

    
