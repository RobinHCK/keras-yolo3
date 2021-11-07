import numpy as np
import configparser
import argparse
from mpl_toolkits.axes_grid1 import host_subplot
import mpl_toolkits.axisartist as AA
from utils import print_config
import sys
import os
import matplotlib.pyplot as plt
from matplotlib import cm
import openpyxl
import xlsxwriter





# =============================================================================
# MAIN
# =============================================================================
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Draw graphics thanks to datas.xlsx')
    parser.add_argument('-c', '--config', type=str, help='The configuration file to use')
    args = parser.parse_args()
    
    # Read the config file given in parameter, if not read the default config file
    if args.config:
        config_path = args.config
    else:
        config_path = 'config.cfg'
        
    if not os.path.isfile(config_path):
        raise ValueError('The configuration file path given does not exist : ', config_path)
        
    config = configparser.RawConfigParser(interpolation = configparser.ExtendedInterpolation())
    config.read(config_path)
    
    
    display_logs = config.getboolean('General', 'display_logs')
    path_logs = config.get('Paths', 'path_logs')

    patients_test = config.get('Train', 'patients_test')
    path_images = config.get('Paths', 'path_images')
    lod = config.getint('MultiScale', 'lod')
    scale_list = config.get('MultiScale', 'scale_list')

    path_dataset_created = config.get('Paths', 'path_dataset_created')
    dataset_name = config.get('Train', 'dataset_name')
    annotations_file_name = config.get('Paths', 'final_annotations_file_name')
    annotation_path = path_dataset_created + dataset_name + '/' + annotations_file_name
    
        
    IOU = config.get('Eval', 'IOU')
    confidenceScore = config.get('Eval', 'confidenceScore')
    drawSurface = config.getboolean('Eval', 'drawSurface')

    tmp = []
    for e in scale_list.split(','):
        tmp.append(int(e))
    scale_list = tmp
    
    tmp = []
    for e in IOU.split(','):
        tmp.append(float(e))
    IOU = tmp
    
    tmp = []
    for e in confidenceScore.split(','):
        tmp.append(float(e))
    confidenceScore = tmp
    
    
    
    sys.stdout = open(path_logs + '/log_DrawGraphics.txt', 'w')
    
    if display_logs :
        print_config(config_path)
    
    for scale in scale_list :
        
        # load datas
        workbook = openpyxl.load_workbook(path_logs + '/datas_' + str(scale) + '.xlsx')
        worksheet = workbook.active 
        
        max_column = worksheet.max_column 
        max_row = worksheet.max_row 
        
        results = []
        result = []
        
        for row in range(1, max_row): 
            result = []
            for column in range(0, max_column):
                result.append(worksheet.cell(row=row+1, column=column+1).value)
            results.append(result)
        
        # draw graphics
        for patient in patients_test.split(','):
            best_r = None
            for r in results:
                if patient == r[8]:
                    if best_r == None:
                        best_r = r
                    else:
                        if r[7] > best_r[7]:
                            best_r = r
                
            print('\n\n\n\nBest F1-Score for patient', best_r[8],'- scale', str(scale),'get with IOU =', best_r[0],'and Confidence score =', best_r[1],'\n')
            print('TP\t\t', best_r[2], '\nFP\t\t', best_r[3], '\nFN\t\t', best_r[4])
            print('precision\t', round(best_r[5],3), '\nrappel\t\t', round(best_r[6],3))
            print('F1 Score\t', round(best_r[7],3))
        
        
        
            # figure 1 : F1 / IOU & Confidence score
            if drawSurface :
                fig = plt.figure()
                ax = fig.add_subplot(111, projection='3d')
                ax.set_xlim3d(0,1)
                ax.set_ylim3d(0,1)
                ax.set_zlim3d(0,1)
                
                X, Y = np.meshgrid(IOU, confidenceScore)
                
                Z = []
                for r in results: 
                    if patient == r[8]:
                        Z.append(r[7])
                        
                Z = np.array(Z).reshape(len(IOU), len(confidenceScore))
                surf = ax.plot_surface(X, Y, Z, vmin=0, vmax=1, cmap=cm.coolwarm, linewidth=0, antialiased=False)
                fig.colorbar(surf, shrink=0.5, aspect=5)
                
                plt.xlabel('Confidence Score')
                plt.ylabel('IOU')
                plt.title('F1-Score evolution for patient ' + patient + ' - scale' + str(scale))
                plt.savefig(path_logs + '/patient_' + patient + '_scale_' + str(scale) + '_F1-Score_IOU_Confidence.png')
    
    
    
            # figure 2 : F1 / IOU 
            fig = plt.figure()
            plt.subplots_adjust(right=0.75)
        
            Y = []
            for r in results: 
                if patient == r[8] and r[1] == best_r[1]:
                    Y.append(r[7])
        
            axes = plt.gca()
            axes.set_xlim([0,1])
            axes.set_ylim([0,1])
            plt.plot(IOU, Y)
            plt.xlabel('IOU')
            plt.ylabel('F1 Score')
            plt.title('F1-Score evolution for patient ' + patient + ' - scale' + str(scale))
            plt.savefig(path_logs + '/patient_' + patient + '_scale_' + str(scale) + '_F1-Score_IOU.png')
            
            
            
            # figure 3 : F1 / Confidence
            Y = []
            Y_TP = []
            Y_FP = []
            Y_FN = []
            for r in results: 
                if patient == r[8] and r[0] == best_r[0]:
                    Y.append(r[7])
                    Y_TP.append(r[2])
                    Y_FP.append(r[3])
                    Y_FN.append(r[4])
    
            fig = plt.figure()
            host = host_subplot(111, axes_class=AA.Axes)
            plt.subplots_adjust(right=0.75)
            
            par2 = host.twinx()
            
            new_fixed_axis = par2.get_grid_helper().new_fixed_axis
            par2.axis["right"] = new_fixed_axis(loc="right", axes=par2,
                                                    offset=(0, 0))
            
            par2.axis["right"].toggle(all=True)
            
            host.set_xlim(0, 1)
            host.set_ylim(0, max(Y_TP))
            
            host.set_xlabel("Confidence")
            host.set_ylabel("Number of TP,FP,FN")
            par2.set_ylabel("F1 Score")
            
            p0, = host.plot(confidenceScore, Y_TP, label="Number of TP")
            p1, = host.plot(confidenceScore, Y_FP, label="Number of FP")
            p2, = host.plot(confidenceScore, Y_FN, label="Number of FN")
            p3, = par2.plot(confidenceScore, Y, label="F1 Score")
            
            par2.set_ylim(0, 1)
            
            host.legend()
    
            par2.axis["right"].label.set_color(p3.get_color())
            
            plt.title('F1-Score evolution for patient ' + patient + ' - scale' + str(scale))
            plt.savefig(path_logs + '/patient_' + patient + '_scale_' + str(scale) + '_F1-Score_TP_FP_FN_Confidence.png')
            
    
    
    # Merge excel files and then compute multiscale evaluation
    # Get results from each scale
    results = []
    for scale in scale_list :
        workbook_scale = openpyxl.load_workbook(path_logs + '/datas_' + str(scale) + '.xlsx')
        worksheet_scale = workbook_scale.active 
        
        max_column = worksheet_scale.max_column 
        max_row = worksheet_scale.max_row 
        
        result = []
        tmp = []
        for i in range(0,max_row):
            for j in range(0,max_column):
                tmp.append(worksheet_scale.cell(row=i+1, column=j+1).value)
            result.append(tmp)
            tmp = []
        
        results.append(result)
            
        workbook_scale.close()
        
    # Merge TP,FP,FN  
    for i in range(1, len(results)) :
        for j in range(1,len(results[i])) :
            for k in range(2,5) :
                results[0][j][k] = results[i][j][k]
    # Update precision, rappel, F1Score 
    for j in range(1,len(results[0])) :
        # precision
        if (results[0][j][2]+results[0][j][3]) == 0 :
            results[0][j][5] = 0
        else :
            results[0][j][5] = results[0][j][2] /(results[0][j][2]+results[0][j][3])

        # rappel
        if (results[0][j][2]+results[0][j][4]) == 0 :
            results[0][j][6] = 0
        else :
            results[0][j][6] = results[0][j][2] /(results[0][j][2]+results[0][j][4])

        # F1Score
        if (results[0][j][5]+results[0][j][6]) == 0 :
            results[0][j][7] = 0
        else :
            results[0][j][7] = (2*results[0][j][5]*results[0][j][6]) / (results[0][j][5]+results[0][j][6])
    
    # Write the result in datas_multiscale.xlsx
    workbook = xlsxwriter.Workbook(path_logs + '/datas_multiscale.xlsx')
    worksheet = workbook.add_worksheet()
        
    for i in range(0,len(results[0])):
        for j in range(0,len(results[0][0])):
            worksheet.write(i, j, results[0][i][j])
        
    workbook.close()
    
    # figure 3 : F1 / Confidence
    workbook = openpyxl.load_workbook(path_logs + '/datas_multiscale.xlsx')
    worksheet = workbook.active 
    
    max_column = worksheet.max_column 
    max_row = worksheet.max_row 
    
    results = []
    result = []
    
    for row in range(1, max_row): 
        result = []
        for column in range(0, max_column):
            result.append(worksheet.cell(row=row+1, column=column+1).value)
        results.append(result)
    
    # draw graphics
    for patient in patients_test.split(','):
        best_r = None
        for r in results:
            if patient == r[8]:
                if best_r == None:
                    best_r = r
                else:
                    if r[7] > best_r[7]:
                        best_r = r
            
        print('\n\n\n\nBest F1-Score for patient', best_r[8],'- multiscale, get with IOU =', best_r[0],'and Confidence score =', best_r[1],'\n')
        print('TP\t\t', best_r[2], '\nFP\t\t', best_r[3], '\nFN\t\t', best_r[4])
        print('precision\t', round(best_r[5],3), '\nrappel\t\t', round(best_r[6],3))
        print('F1 Score\t', round(best_r[7],3))
                
        Y = []
        Y_TP = []
        Y_FP = []
        Y_FN = []
        for r in results: 
            if patient == r[8] and r[0] == best_r[0]:
                Y.append(r[7])
                Y_TP.append(r[2])
                Y_FP.append(r[3])
                Y_FN.append(r[4])
    
        fig = plt.figure()
        host = host_subplot(111, axes_class=AA.Axes)
        plt.subplots_adjust(right=0.75)
        
        par2 = host.twinx()
        
        new_fixed_axis = par2.get_grid_helper().new_fixed_axis
        par2.axis["right"] = new_fixed_axis(loc="right", axes=par2,
                                                offset=(0, 0))
        
        par2.axis["right"].toggle(all=True)
        
        host.set_xlim(0, 1)
        host.set_ylim(0, max(Y_TP))
        
        host.set_xlabel("Confidence")
        host.set_ylabel("Number of TP,FP,FN")
        par2.set_ylabel("F1 Score")
        
        p0, = host.plot(confidenceScore, Y_TP, label="Number of TP")
        p1, = host.plot(confidenceScore, Y_FP, label="Number of FP")
        p2, = host.plot(confidenceScore, Y_FN, label="Number of FN")
        p3, = par2.plot(confidenceScore, Y, label="F1 Score")
        
        par2.set_ylim(0, 1)
        
        host.legend()
    
        par2.axis["right"].label.set_color(p3.get_color())
        
        plt.title('F1-Score evolution for patient ' + patient + ' - multiscale')
        plt.savefig(path_logs + '/patient_' + patient + '_multiscale_F1-Score_TP_FP_FN_Confidence.png')


